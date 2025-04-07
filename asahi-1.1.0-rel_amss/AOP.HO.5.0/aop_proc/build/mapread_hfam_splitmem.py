# vim: set ts=4 sw=4 noet smarttab ai :

import os, re, sys, subprocess, csv, argparse
from itertools import islice, count


# dictionary of dictionaries representing image sizes
# first dictionary is keyed on reference build, e.g. 30, 31, 32
# second dictionary is keyed on "driver" e.g. sleep, kernel, etc.
#library_sizes = { }

library_map = {
    'glink_api.*'                       : 'mproc',
    'glink_os_rex.o'                    : 'mproc',
    'xport_qmp*'                        : 'mproc',
    'startup.o'                         : 'aop_startup',
    'main.o'                            : 'aop_main',
    'target_main.o'                     : 'aop_main',
    'retarget.o'                        : 'aop_retarget',
    'exceptions.o'                      : 'aop_exception',
    'time_service.o'                    : 'aop_time_service',
    'coredump.o'                        : 'aop_coredump',
    'stack_protect.o'                   : 'aop_stack_gap',
    'aop_interrupt_table.o'             : 'aop_interrupt_table',
    'bcm_init_vals'                     : 'bcm',
    'rex.*.lib'                         : 'rex',
    'pmic_framework.lib'                : 'pmic',
    'pmic_target_rpm.lib'               : 'pmic',
    'pmic_npa.lib'                      : 'pmic',
    'pmic_[^\s]+.lib'                   : 'pmic',
    'vrm.lib'                           : 'vrm',
    'arc.lib'                           : 'arc',
    'arc_init_vals'                     : 'arc',
    'arc_init'                          : 'arc',
    'cmd_db.lib'                        : 'cmd_db',
    'cpr.lib'                           : 'cpr',
    'cpr_init_vals'                     : 'cpr',
    'cpr_init.*'                        : 'cpr',
    'ddr_mgr_arc.lib'                   : 'ddr_mgr',
    'ddraux_init_vals'                  : 'ddr_mgr',
    'ddr_mgr_init.lib'                  : 'ddr_mgr',
    'pdc[^\s]+.lib'                     : 'pdc',
    'pdc_global_init_vals'              : 'pdc',
    'rinit.lib'                         : 'rinit',
    'rpmh.lib'                          : 'rpmh',
    'sleep.lib'                         : 'sleep',
    'sleep_.*.lib'                      : 'sleep',
    'ulog.lib'                          : 'ulog_driver',
    'pwr_utils.lib'                     : 'pwr_utils',
    'icb.*'                             : 'bcm',
    'bcm_starc_.*'                      : 'bcm',
    'Spmi*'                             : 'spmi',
    'Clock.lib'                         : 'clock',
    'clock_init_vals'                   : 'clock',
    'PlatformInfo.lib'                  : 'sysdrivers',
    'aop_fsm_arc.lib'                   : 'aop_fsm',
    'aop_fsm_init.lib'                  : 'aop_fsm',
    'rsc.lib'                           : 'rsc',
    'rsc_init.*'                        : 'rsc',
    'libc.a'                            : 'libc',
    'oem_version.o'                     : 'aop_oem',
    'qc_version.o'                      : 'aop_oem',
    'datetime.o'                        : 'aop_oem',
    'exception_isr.o'                   : 'aop_exception',
    'oem_uuid.o'                        : 'aop_oem',
    'crt1.o'                            : 'aop',
    'aop_mpu*'                          : 'mpu',
    'unpa*'                             : 'unpa',
    'qdss*'                             : 'qdss',
    'ChipInfo.lib'                      : 'chip_info',
    'PM_CRAM_RECLAIM_POOL'              : 'PM_CRAM_RECLAIM',
    'DDR_ULOG'                          : 'ulog',
    'PM_DRAM_RECLAIM_POOL'              : 'PM_DRAM_RECLAIM',
    'cprv'                              : 'cpr',
    'bcm_v'                             : 'bcm',
    'bcm_db_headers'                    : 'bcm',
    'arc_v'                             : 'arc',
    'arc_db_headers'                    : 'arc',
    'arc_aux'                           : 'arc',
    'arc_mol_aux'                       : 'arc',
    'PMIC_DATA'                         : 'PMIC_DATA',
    'bcm_aux'                           : 'bcm',
    'cprwa_init_vals'                   : 'cpr',
    'vrm_tcs_no_pmi_init_vals'          : 'vrm',
    'vrm_tcs_init_vals'                 : 'vrm',
    'vrm_init_vals'                     : 'vrm',
    'IPCC*'                             : 'IPCC',
    'audio_pdc_init_vals'               : 'pdc',
    'sensors_pdc_init_vals'             : 'pdc',
    'compute_pdc_init_vals'             : 'pdc',
    'pdc_seq_init_vals'                 : 'pdc',
    'wlan_pdc_s'                        : 'pdc',
    'limits_init_vals'                  : 'limits',
    'pwr_utils_init'                    : 'pwr_utils',
    'mi_pmu'                            : 'bcm',
    'mi_pmu_init_vals'                  : 'bcm',
    'pdc_tcs_init_vals'                 : 'pdc',
    'vrm_init*'                         : 'vrm',
    'AOP_SDI'                           : 'AOP_SDI',
    'settings*'                         : 'devcfg',
    'aop_accesscontrol'                 : 'access control'
}

section_map = {
    'RINIT_DATA'                        : 'RINIT DDR DATA',
    'CODE_RAM'                          : 'CODE RAM',
    'ARM_LIB_HEAP'                      : 'LIB HEAP',
    'ARM_LIB_STACK'                     : 'LIB STACK',
    'DATA_RAM'                          : 'DATA RAM',
    'AOP_ULOG'                          : 'AOP ULOG',
    'DDR_ULOG'                          : 'DDR ULOG',
    'PMIC_DATA'                         : 'PMIC DATA DDR',
    'TASK_STACKS'                       : 'TASK STACKS',
    'PM_DRAM_RECLAIM_POOL'              : 'DRAM_RECLAIM',
    'PM_CRAM_RECLAIM_POOL'              : 'CRAM_RECLAIM',
    'AOPSS_DATA'                        : 'DATA RAM',
    'DRAM_SLEEP_CODE_POOL'              : 'DRAM_SLEEP_CODE_POOL',
    'DDR_CODE'                          : 'DDR OFFLOADED CODE-DATA',
    'CPR_CRAM_RECLAIM_POOL'             : 'CRAM_RECLAIM',
    'SYSDBG_CODE'                       : 'SYSTEM DEBUG',
    'SYSDBG_DUMMY_ZI'                   : 'SYSTEM DEBUG DUMMY 4K',
    'AOP_CRAM_ZI'                       : 'CODE RAM ZI',
    'AOP_DRAM_ZI'                       : 'DRAM RAM ZI',
    'AOPSS_PMIC_ZI'                     : 'AOPSS PMIC ZI',
    'AOPSS_DDR_CODE_ZI'                 : 'AOPSS DDR ZI',
     
}

MEMORY_BREAKDOWN = {
  'CRAM' : {'SECTIONS' : [
                          'CODE_RAM',
                          'AOP_ULOG',
                          'DDR_ULOG',
                          'ARM_LIB_HEAP',
                          'ARM_LIB_STACK',
                          'PM_CRAM_RECLAIM_POOL',
                          'CPR_CRAM_RECLAIM_POOL'
                          ],
            'SIZE': 0x18000,
          },
  'DRAM' : {'SECTIONS' : [
                          'DATA_RAM',
                          'TASK_STACKS',
                          'AOPSS_DATA',
                          ],
            'SIZE': 0x8000,
          },
  'DDR' : {'SECTIONS' : [
                          'RINIT_DATA',
                          'PMIC_DATA',
                          'DDR_CODE',
                          ],
            'SIZE': 0x18000,
          },
}



current_section = ""
found_end_regex = False

def get_lib_line_data(m):
    base_addr = int(m.group('base_addr') , base=16)
    size = int(m.group('size') , base=16)
    lib = m.group('lib').strip()
    if '(' in lib: # #bracket libnames not valid
        split_libname = re.split('\(|\/', lib) # hack to get obj file name 
        lib = split_libname[len(split_libname)-2]
    return lib, base_addr, size


def get_sec_line_data(m):
    size = int(m.group('size'), base=16)
    base_addr = int(m.group('base_addr'), base=16)
    section = m.group('section').strip()
    return section , base_addr, size

def account_align_padding(m):
    size = int(m.group('size'), base=16)
    return size

def cram_space_left(m):
    return int(m.group('size_left'), base=16)

map_line_types_dict = {
    #dictionary of regexs for the type of lines expected and their corresponding parser functions
    'lib' : {'regex': re.compile(r'(?P<subsec>[_.a-zA-Z0-9]+)\s*(?P<base_addr>0x[0-9a-fA-F]+)\s+(?P<size>0x[0-9a-fA-F]+)\s+(?P<path>[^\s]*[\/|\\])(?P<lib>[^\s]*\.o)'),
                  'parser' : get_lib_line_data},
    'section' : {'regex' : re.compile(r'(?P<section>[_0-9a-zA-Z]*)\s+(?P<base_addr>0x[0-9a-fA-F]+)\s+(?P<size>0x[0-9a-fA-F]+)\s*\#\s*[^\s*]*\s*(?P<offset_addr>0x[0-9a-fA-F]+)\,\s+[LMA: ]+(?P<lma_addr>0x[0-9a-fA-F]+)'),
                  'parser' : get_sec_line_data},
    'end'  :     {'regex' : re.compile(r'\.ARM\.attributes'),
                 'parser' : None},
    'padding': {'regex': re.compile(r'PADDING_ALIGNMENT\s+(?P<base_addr>0x[0-9a-fA-F]+)\s+(?P<size>0x[0-9a-fA-F]+)'),
                'parser' : account_align_padding,
               },
    'cram_size_left' :{'regex': re.compile(r'.+__aop_cram_size_left\((?P<size_left>\s+0x[0-9a-fA-F]+)'),
                      'parser': cram_space_left},
}


def get_module_pretty_name(module):
    # figure out if we should attribute this to a renamed module, or to it's own name
    for pattern in library_map:
        m = pattern.match(module)
        if m:
            return library_map[pattern]

    #print >> sys.stderr, "*** WARNING: no pretty renaming for module %s" % (module)
    return module

    
def get_section_pretty_name(section):
    for pattern in section_map:
        m = pattern.match(section)
        if m:
            return section_map[pattern]

    print("*** WARNING: no pretty renaming for section %s" % (section))
    return section
    
def _debug_print(library_sizes):
  for sections in library_sizes:
    print (sections + ': \n')
    for libs in library_sizes[sections]:
      print ("\t" + libs + " : " + str(library_sizes[sections][libs]))

def dump_xl(library_sizes):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Series, Reference

    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'AOP Memory breakdown'

    section_list = []
    header_row = 1
    lib_start_row = 2
    lib_curr_row = lib_start_row
    for lib in library_sizes:
        #print lib, '<-- lib'
        libcell = ws.cell(column=1, row=lib_curr_row, value="{0}".format(lib))
        libcell.fill = PatternFill("solid", fgColor="DDDDDD")
        for section in library_sizes[lib]:
            if section not in section_list: 
                section_list.append(section)
                section_cell = ws.cell(column=section_list.index(section)+2, row=header_row, value="{0}".format(section))
                ws.column_dimensions[get_column_letter(section_list.index(section)+2)].width = len(section) + 10
                section_cell.fill = PatternFill("solid", fgColor="7e86de")
            ws.cell(column=section_list.index(section)+2, row = lib_curr_row, value=(library_sizes[lib][section]))
        #next lib on new line
        lib_curr_row = lib_curr_row + 1

    #section sums
    lib_end_row = lib_curr_row
    ws.cell(column=1, row=lib_end_row, value="SUM").fill = PatternFill("solid", fgColor="F1948A") 
    for section in section_list:
      col_let = get_column_letter(section_list.index(section)+2)
      sum_cells = ws.cell(column=section_list.index(section)+2, row=lib_end_row, value="=SUM(${0}${1}:${0}${2})".format(col_let,lib_start_row,lib_end_row-1))
      sum_cells.border = border
      sum_cells.fill = PatternFill("solid", fgColor="FADBD8")
  
    #lib sums
    col = len(section_list)+2
    ws.cell(column=col, row=header_row, value="SUM").fill = PatternFill("solid", fgColor="F1948A")
    for row in range(header_row+1, lib_end_row):
      sum_cells = ws.cell(column=col, row=row, value="=SUM(${0}${1}:${2}${1})".format(get_column_letter(2),row,get_column_letter(col-1)))
      sum_cells.border = border
      sum_cells.fill = PatternFill("solid", fgColor="FADBD8")

    for row in range(1,lib_end_row):
        for col in range(1, len(section_list)+2):
            ws.cell(column=col, row=row).border = border

    bchart = BarChart()
    bchart.type = "col"
    bchart.style = 10
    bchart.title = "AOP Memsizes"
    bchart.height = 20
    bchart.width = 90
    bchart.y_axis.title = 'Memsize'
    bchart.x_axis.title = 'Libs'

    for col in range(2, len(section_list)+2):  
      if section_list[col-2] in ['CODE RAM', 'DATA RAM', 'CRAM_RECLAIM', 'DDR OFFLOADED CODE-DATA']:
        #data = Reference(ws, min_col=2, max_col=len(section_list)+1, min_row=header_row, max_row=lib_end_row-1)
        data = Reference(ws, min_col=col, min_row=header_row, max_row=lib_end_row-1)
        bchart.add_data(data, titles_from_data=True)

    libs = Reference(ws, min_col=1, min_row=header_row+1, max_row=lib_end_row)

    #bchart.add_data(data, titles_from_data=True)
    bchart.set_categories(libs)
    bchart.shape = 4
    ws_chart = wb.create_sheet("Analysis")
    ws_chart.add_chart(bchart, "A1")

    wb.save(filename = 'memsizes.xlsx')

def dump_csv(library_sizes):
         
    # Create a local CSV file
    #_debug_print()
    header_list = []
    data_list   = []
    map_path = library_sizes['path']
    del library_sizes['path']
    for libs in library_sizes:
        agg_lib_size = 0
        for sections in library_sizes[libs]:
            header_list.append('%s:%s' % (libs, sections))
            #print sections
            if(sections != 'RINIT DATA'):
                agg_lib_size += library_sizes[libs][sections]
            data_list.append('%s' % library_sizes[libs][sections])
        header_list.append('%s: Total' % libs)
        data_list.append('%s' % agg_lib_size) 
  
    #print map_path
    with open('memsizes.csv', "w", encoding='utf-8') as opfile:
        wr = csv.writer(opfile, dialect = 'excel')
        wr.writerow(header_list)
        wr.writerow(data_list)
    
def parse_each_line(line):
  for line_type in map_line_types_dict:
    matches = map_line_types_dict[line_type]['regex'].match(line)
    if matches and map_line_types_dict[line_type]['parser'] is not None:
      parsed_line = map_line_types_dict[line_type]['parser'](matches)
      return line_type, parsed_line
  return None, None

def is_end_line(line):
  if map_line_types_dict['end']['regex'].match(line):
    return True
  else:
    return False

def add_or_increment_lib_size(lib_sizes, lib, size):
  curr_section = lib_sizes['current_section']
  if lib not in lib_sizes:
    lib_sizes[lib] = {}
  if curr_section not in lib_sizes[lib]:
    lib_sizes[lib][curr_section] = size
  else:
    lib_sizes[lib][curr_section] += size

#curr_section = ''
def build_lib_map(lib_sizes, line_type, parsed):
  if line_type == 'section':
    #mustve seen a new section, update curr section
    lib_sizes['current_section'] = get_section_pretty_name(parsed[0])
    if lib_sizes['current_section'] in ['LIB HEAP', 'LIB STACK', 'AOP ULOG', 'DDR ULOG']:
      #common sections
      add_or_increment_lib_size(lib_sizes, 'common', parsed[2])
  if line_type == 'lib':
    add_or_increment_lib_size(lib_sizes, get_module_pretty_name(parsed[0]), parsed[2])
  if line_type == 'padding':
    add_or_increment_lib_size(lib_sizes, 'PAD', parsed)

def get_map_lines(map_file):
    #empty dict for this map file
    lib_sizes = {}
    lib_sizes['current_section'] = ''
    m = re.match(r'(.*)aop_proc', map_file)
    if m: lib_sizes['path'] = m.group(1) 
    else: lib_sizes['path'] = map_file
    with open(map_file) as mapfile: lines = mapfile.readlines()
    for line in lines:
      if not is_end_line(line):
        (line_type, parsed) = parse_each_line(line)
        if line_type is not None: build_lib_map(lib_sizes, line_type, parsed)
      else:
        #print 'done parsing'
        del lib_sizes['current_section']
        return lib_sizes
    
if __name__ == '__main__':
    # regenerate the pretty library renaming map with actual compiled regexs
    library_map_compiled = {}
    section_map_compiled = {}
    for regex in library_map:
        library_map_compiled[re.compile(regex)] = library_map[regex]
    library_map = library_map_compiled

    for regex in section_map:
        section_map_compiled[re.compile(regex)] = section_map[regex]
    section_map = section_map_compiled

    parser = argparse.ArgumentParser()
    #parser.add_argument('-del', '--delta', help='size delta, comma separate map file paths', action='store_true', dest='delta_req', required=False)
    parser.add_argument('-f', '--file', help='path to map file', metavar='FILE', dest='filename', required=False)
    parser.add_argument('-dcsv', '--dump_csv', action='store_true', help='Dump csv?', dest='dump_csv', default=True, required=False)
    parser.add_argument('-dxl', '--dump_xl', action='store_true', help='Dump excel?', dest='dump_excel', default=False, required=False)
    parser.add_argument('-dr', '--dump_raw', action='store_true', help='Dump raw?', dest='dump_raw', default=False, required=False)
    args = parser.parse_args()

    map_file_locations = [
          r'aop_proc/core/bsp/aop/build/'+os.environ.get('MSM_ID', '8350')+'/AOP_' + os.environ.get('BUILD_ID', 'AAAAANAAO') + '.map',
    ]

    if os.environ.get('BIN_FOLDER_NAME') is not None: 
        map_file_locations[0] = r'aop_proc/core/bsp/aop/build/'+os.environ['BIN_FOLDER_NAME']+'/AOP_' + os.environ.get('BUILD_ID', 'AAAAANAAO') + '.map'

    mapfilepaths = []
    # arg parsing
    if args.filename:
        #file name given use that
        mapfiles = args.filename
        #file name can be comma seperate for delta
        in_map_files = [mpath.strip(" ") for mpath in mapfiles.split(",")]
        
        for mapfile in in_map_files:
          if os.path.exists(os.path.abspath(mapfile)):
            mapfilepaths.append(os.path.abspath(mapfile))
    else:
        if "aop_proc" in os.getcwd():
            #we are executing in build path proceed to find map file
            m = re.match(r'(.*)aop_proc', os.getcwd())
            builds = m.group(1)
            if os.path.exists(os.path.abspath(os.path.join(builds, map_file_locations[0]))):
              mapfilepaths.append(os.path.abspath(os.path.join(builds, map_file_locations[0])))
            else:
              print ("can't find path, " ,os.path.join(builds, map_file_locations[0]))
              sys.exit(1)

    for mapfilepath in mapfilepaths:
      if os.path.isfile(mapfilepath):
        library_sizes_map = get_map_lines(mapfilepath)
        if args.dump_csv:
            dump_csv(library_sizes_map)
            print ('Created memsizes.csv in local dir')
        if args.dump_raw:
          _debug_print(library_sizes_map)
        if args.dump_excel:
          dump_xl(library_sizes_map)
          print ('Created memsizes.xlsx in local dir')
      else:
          print ("can't find path, " ,mapfilepath)
          sys.exit(1)
